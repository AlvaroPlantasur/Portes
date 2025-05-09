import os
import psycopg2
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from datetime import datetime
from dateutil.relativedelta import relativedelta
import sys
import copy

def main():
    # 1. Obtener credenciales y ruta del archivo
    db_name = os.environ.get('DB_NAME')
    db_user = os.environ.get('DB_USER')
    db_password = os.environ.get('DB_PASSWORD')
    db_host = os.environ.get('DB_HOST')
    db_port = os.environ.get('DB_PORT')
    file_path = os.environ.get('EXCEL_FILE_PATH')
    
    db_params = {
        'dbname': db_name,
        'user': db_user,
        'password': db_password,
        'host': db_host,
        'port': db_port
    }
    
     # 2. Definir la nueva consulta SQL con fechas dinámicas
    fecha_inicio_str = '2025-01-01'
    fecha_fin = datetime.now().date()
    fecha_fin_str = fecha_fin.strftime('%Y-%m-%d')
    
    # 3. Consulta SQL con el rango dinámico
    query = f"""
    SELECT 
    ai.id AS "ID FACTURA",
    ai.date_invoice AS "FECHA FACTURA",
    ai.internal_number AS "CODIGO FACTURA",
    ai.name AS "DESCRIPCION",
    rc.name AS "COMPAÑÍA",
    ssp.name AS "SEDE",
    'S-' || rp.id AS "ID BBSeeds",
    rp.nombre_comercial AS "CLIENTE",
    rpa.city AS "CIUDAD",
    CASE 
        WHEN rpa.prov_id IS NOT NULL THEN 
            (SELECT UPPER(name) FROM res_country_provincia WHERE id = rpa.prov_id)
        ELSE 
            UPPER(rpa.state_id_2)
    END AS "PROVINCIA",
    CASE 
        WHEN rpa.cautonoma_id IS NOT NULL THEN 
            (SELECT UPPER(name) FROM res_country_ca WHERE id = rpa.cautonoma_id)
        ELSE 
            ''
    END AS "COMUNIDAD",
    c.name AS "PAÍS",
    EXTRACT(MONTH FROM ai.date_invoice) AS "MES",
	EXTRACT(DAY FROM ai.date_invoice) AS "DÍA",
    CASE 
        WHEN ai.type = 'out_invoice' THEN 
            COALESCE(ai.portes, 0) + COALESCE(ai.portes_cubiertos, 0)
        ELSE 
            -(COALESCE(ai.portes, 0) + COALESCE(ai.portes_cubiertos, 0))
    END AS "PORTES CARGADOS POR EL TRANSPORTISTA",
    CASE 
        WHEN ai.type = 'out_invoice' THEN 
            COALESCE(ai.portes_cubiertos, 0)
        ELSE 
            -(COALESCE(ai.portes_cubiertos, 0))
    END AS "PORTES CUBIERTOS",
    CASE 
        WHEN ai.type = 'out_invoice' THEN 
            COALESCE(ai.portes, 0)
        ELSE 
            -(COALESCE(ai.portes, 0))
    END AS "PORTES COBRADOS A CLIENTE",
    EXTRACT(YEAR FROM ai.date_invoice) AS "AÑO"
 
FROM 
    account_invoice ai
INNER JOIN 
    res_partner_address rpa ON rpa.id = ai.address_shipping_id
INNER JOIN 
    res_country c ON c.id = rpa.pais_id
LEFT JOIN 
    stock_sede_ps ssp ON ssp.id = ai.sede_id
LEFT JOIN 
    res_company rc ON rc.id = ai.company_id
LEFT JOIN 
    res_partner rp ON rp.id = ai.partner_id
 
WHERE 
    ai.state NOT IN ('draft', 'cancel') 
    AND ai.type IN ('out_invoice', 'out_refund') 
    AND ai.carrier_id IS NOT NULL 
    AND ai.date_invoice BETWEEN '{fecha_inicio_str}' AND '{fecha_fin_str}' 
    AND ai.obsolescencia = FALSE -- BBSEEDS
 
GROUP BY 
    ai.id,
    ai.company_id,
    ai.date_invoice,
    EXTRACT(YEAR FROM ai.date_invoice),
    EXTRACT(MONTH FROM ai.date_invoice),
    EXTRACT(DAY FROM ai.date_invoice),
    ai.carrier_id,
    ai.partner_id,
    ai.name,
    ai.obsolescencia,
    ai.type,
    c.name,
    rpa.state_id_2,
    rc.name,
    ssp.name,
    rpa.prov_id,
    rpa.cautonoma_id,
    rp.nombre_comercial,
    rpa.city,
    rp.id
 
ORDER BY 
    ai.id DESC;
    """
    
    # 4. Ejecutar la consulta
    try:
        with psycopg2.connect(**db_params) as conn:
            with conn.cursor() as cur:
                cur.execute(query)
                resultados = cur.fetchall()
                headers = [desc[0] for desc in cur.description]
    except Exception as e:
        print(f"Error al conectar o ejecutar la consulta: {e}")
        sys.exit(1)
    
    if not resultados:
        print("No se obtuvieron resultados de la consulta.")
        return
    else:
        print(f"Se obtuvieron {len(resultados)} filas de la consulta.")
    
    # 5. Cargar el archivo base Portes.xlsx que se encuentra en la raíz del repositorio
    try:
        book = load_workbook(file_path)
        sheet = book.active
    except FileNotFoundError:
        print(f"No se encontró el archivo base '{file_path}'. Se aborta para no perder el formato.")
        return
    
    # 6. Evitar duplicados (usando la tercera columna: "CODIGO FACTURA")
	existing_invoice_codes = {row[2] for row in sheet.iter_rows(min_row=2, values_only=True) if row[2] is not None}
    for row in resultados:
    	if row[2] not in existing_invoice_codes:
        sheet.append(row)
        new_row_index = sheet.max_row
        if new_row_index > 1:
            for col in range(1, sheet.max_column + 1):
                source_cell = sheet.cell(row=new_row_index - 1, column=col)
                target_cell = sheet.cell(row=new_row_index, column=col)
                target_cell.font = copy.copy(source_cell.font)
                target_cell.fill = copy.copy(source_cell.fill)
                target_cell.border = copy.copy(source_cell.border)
                target_cell.alignment = copy.copy(source_cell.alignment)

    
    # 7. Actualizar la referencia de la tabla existente
    # Asumimos que la tabla se llama "Portes"
    if "Portes" in sheet.tables:
        tabla = sheet.tables["Portes"]
        max_row = sheet.max_row
        max_col = sheet.max_column
        last_col_letter = get_column_letter(max_col)
        new_ref = f"A1:{last_col_letter}{max_row}"
        tabla.ref = new_ref
        print(f"Tabla 'Portes' actualizada a rango: {new_ref}")
    else:
        print("No se encontró la tabla 'Portes'. Se conservará el formato actual, pero no se actualizará la referencia de la tabla.")
    
    book.save(file_path)
    print(f"Archivo guardado con la estructura de tabla en '{file_path}'.")
    
if __name__ == '__main__':
    main()
